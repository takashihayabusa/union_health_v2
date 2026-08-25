from django.contrib.auth.hashers import check_password
from .forms import AccountRegisterForm, LoginForm
from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse, FileResponse, Http404
from django.contrib.auth.hashers import make_password
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Font, PatternFill, Alignment

from linebot import LineBotApi, WebhookHandler
from linebot.models import *

from django.conf import settings
from .models import LineUser, LineLog, Account, BroadcastHistory

from openpyxl.styles import PatternFill, Font
from datetime import datetime
import re
from .services import send_health_check_to_all
from pathlib import Path
from django.core import signing
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile


print("CHECKAPP 起動")


from .line_api import (
    line_bot_api,
    handler,
)
# =====================================
# LINE LOG 共通保存
# =====================================

def save_line_log(user, sender, message_type, content):

    try:

        LineLog.objects.create(
            user=user,
            sender=sender,
            message_type=message_type,
            content=content
        )

    except Exception as e:

        print("LINE LOG 保存エラー:", e)


# -------------------------
# 表紙
# -------------------------
from django.shortcuts import redirect

def home(request):
    return render(request, "checkapp/home.html")

def union_home(request):

    if "account_id" not in request.session:
        return redirect("login")

    return render(request, "checkapp/union_home.html")

# -------------------------
# ユーザー一覧
# -------------------------
def user_list(request):

    users = (
        LineUser.objects
        .all()
        .order_by("-created_at")
    )

    # GETのときは新しい登録画面を表示

    return render(
        request,
        "checkapp/account_register.html",
        {
        "form": form,
        "user_id": user_id,
        }
)
# -------------------------
# 登録
# -------------------------
def register(request):

    # LINEのuser_idを取得
    user_id = request.GET.get("user_id") or request.POST.get("user_id")

    if request.method == "POST":

        form = AccountRegisterForm(request.POST)

        if form.is_valid():

            name = form.cleaned_data["name"]
            region = form.cleaned_data["region"]
            login_id = form.cleaned_data["login_id"]
            birth_date = form.cleaned_data["birth_date"]
            password = form.cleaned_data["password1"]

            # 組合員アカウントを保存
            Account.objects.create(
                login_id=login_id,
                password=make_password(password),
                name=name,
                area=region,
                birth_date=birth_date,
            )

            # LINE利用者情報を保存
            if user_id:
                LineUser.objects.update_or_create(
                    user_id=user_id,
                    defaults={
                        "login_id": login_id,
                        "name": name,
                        "region": region,
                    }
                )

            # 登録完了画面へ
            return render(
                request,
                "checkapp/register_complete.html"
            )

    else:
        form = AccountRegisterForm()

    return render(
        request,
        "checkapp/account_register.html",
        {
            "form": form,
            "user_id": user_id,
        }
    )
# ==========================
# 地域選択画面
# ==========================
def region_send(request):

    return render(
        request,
        "checkapp/region_send.html"
    )
# -------------------------
# 健康チェック送信
# -------------------------
from .services import send_health_check_to_all

def send_health_check(request):

    success, error = send_health_check_to_all()

    return HttpResponse(
        f"健康チェック送信完了 成功:{success}件 失敗:{error}件"
    )
# -------------------------
# 緊急生存確認
# -------------------------
users = LineUser.objects.all()
# -------------------------
# Callback
# -------------------------
@csrf_exempt
def callback(request):

    body = request.body.decode("utf-8")
    signature = request.META.get("HTTP_X_LINE_SIGNATURE", "")

    try:
        handler.handle(body, signature)

    except Exception as e:
        print("Callback Error:", e)

    return HttpResponse("OK")
@handler.add(FollowEvent)
def handle_follow(event):

    user_id = event.source.user_id

    LineUser.objects.get_or_create(
        user_id=user_id
    )

    register_url = (
        f"https://nonfrigid-smug-candance.ngrok-free.dev/register?user_id={user_id}"
    )

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(
            text=(
                "友だち追加ありがとうございます。\n\n"
                "マルキョウユニオンLINEをご利用いただくため、"
                "組合員登録をお願いします。\n\n"
                "▼ 組合員登録はこちら\n"
                f"{register_url}"
            )
        )
    )
# -------------------------
# メッセージ受信
# -------------------------
@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):

    user_id = event.source.user_id
    
    text = event.message.text

    user, created = LineUser.objects.get_or_create(
        user_id=user_id
        )

    print("受信メッセージ:", text)
    print("現在のstep:", user.step)

    if text == "元気です":

        reply = TextSendMessage(
            text="それは何よりです。今日も一日頑張りましょう。"
        )

        LineLog.objects.create(
            user=user,
            sender="BOT",
            message_type="text",
            content="それは何よりです。今日も一日頑張りましょう。"
        )
        line_bot_api.reply_message(
        event.reply_token,
        reply
)

    elif text == "少し疲れています":

        reply = TextSendMessage(
            text="無理をしないでください。休息を取ってください。"
        )
        line_bot_api.reply_message(
        event.reply_token,
        reply
        )

    elif text == "かなり辛いです":

        reply = TextSendMessage(
            text="その原因は体調ですか？ メンタルですか？",
            quick_reply=QuickReply(
                items=[
                    QuickReplyButton(
                        action=MessageAction(label="体調", text="体調")
                    ),
                    QuickReplyButton(
                        action=MessageAction(label="メンタル", text="メンタル")
                    ),
                ]
            )
        )
        
        line_bot_api.reply_message(
        event.reply_token,
        reply
        )

    elif text == "体調":

        reply = TextSendMessage(
            text="無理をせず、病院の受診も検討してください。お大事になさってください。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )

    elif text == "メンタル":

        reply = TextSendMessage(
            text="原因を教えてください。",
            quick_reply=QuickReply(
                items=[
                    QuickReplyButton(
                        action=MessageAction(
                            label="個人的な悩み",
                            text="個人的な悩み"
                        )
                    ),
                    QuickReplyButton(
                        action=MessageAction(
                            label="仕事の悩み",
                            text="仕事の悩み"
                        )
                    ),
                ]
            )
        )

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )

    elif text == "個人的な悩み":

        reply = TextSendMessage(
            text="一人で抱え込まず、信頼できる方へ相談してください。"
        )

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )

    elif text == "仕事の悩み":

        reply = TextSendMessage(
        text="委員長に相談できます。\nどうしますか？",
        quick_reply=QuickReply(
            items=[
                QuickReplyButton(
                    action=MessageAction(
                        label="お願いします",
                        text="お願いします"
                    )
                ),
                QuickReplyButton(
                    action=MessageAction(
                        label="いいえ結構です",
                        text="いいえ結構です"
                    )
                ),
            ]
        )
    )
        line_bot_api.reply_message(
            event.reply_token,
            reply
            )
        return
    elif text == "お願いします":
        user.step = ""
        user.save()

        reply = TextSendMessage(
            text="相談方法を選んでください。",
            quick_reply=QuickReply(
                items=[
                    QuickReplyButton(
                        action=MessageAction(
                            label="☎️ 電話",
                            text="☎️ 電話"
                        )
                    ),
                    QuickReplyButton(
                        action=MessageAction(
                            label="LINE",
                            text="LINEで相談"
                        )
                    ),
                ]
            )
        )

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )
        return

    elif text == "☎️ 電話":

        reply = TextSendMessage(
            text=(
                "委員長へ電話で相談できます。\\n\\n"
                "☎️ 090-9495-5990"
            )
        )

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )
        return

    elif text == "LINEで相談":

        image_url = (
            "https://nonfrigid-smug-candance.ngrok-free.dev"
            "/static/images/kumaki.jpg"
        )

        reply = [
            TextSendMessage(
                text="委員長のLINEをQRコードから登録してください。"
            ),
            ImageSendMessage(
                original_content_url=image_url,
                preview_image_url=image_url
            ),
        ]

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )
        return
    elif user.step == "wait_phone_number":

    # USERログ
        save_line_log(
            user,
            "USER",
            "電話番号",
            text
            )

        user.step = ""
        user.save()

        reply = TextSendMessage(
            text=(
                "電話番号ありがとうございました。\n\n"
                "委員長へ連絡いたします。"
            )
        )

    # BOTログ
        save_line_log(
            user,
            "BOT",
            "健康チェック",
            "電話番号ありがとうございました。委員長から連絡させます。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )
        return

    elif user.step == "wait_emergency_phone":

    # USERログ
        save_line_log(
            user,
            "USER",
            "緊急電話番号",
            text
            )

        user.step = ""
        user.save()

        reply = TextSendMessage(
        text=(
            "電話番号を確認しました。\n\n"
            "ありがとうございます。\n\n"
            "現在の状況を確認し、救援や電話での対応ができないか確認いたします。\n\n"
            "今の状況が変ったら下記に電話ください\n\n"
            "【委員長】\n"
            "TEL：080-1720-0311\n\n"
            "【組合】\n"
            "TEL：092-513-9820\n\n"
            )
        )

    # BOTログ
        save_line_log(
            user,
            "BOT",
            "緊急",
            "電話番号を確認しました。現在の状況を確認し、救援や電話での対応ができないか確認いたします。委員長または組合からご連絡いたしますので、しばらくお待ちください。状況が変わった場合や、安全な場所へ避難できた場合は、そのままLINEでお知らせください。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )
        return

    # 電話番号を保存
    # USERログ
        save_line_log( 
            user,
            "USER",
            "電話番号",
            text
            )

        reply = TextSendMessage(
            text=(
                "電話番号を確認しました。\n\n"
                "ありがとうございます。\n\n"
                "委員長または組合からご連絡いたしますので、\n"
                "しばらくお待ちください。"
            )
)

    # BOTログ
        save_line_log(
            user,
            "BOT",
            "電話番号",
            "電話番号を確認しました。委員長または組合からご連絡いたしますので、しばらくお待ちください。"
)

        line_bot_api.reply_message(
            event.reply_token,
            reply
        )
        return
    


    elif text == "いいえ結構です":

        user.step = ""
        user.save()

        reply = TextSendMessage(
            text=(
                "会社には内部通報がありますので\n"
                "ご利用ください。\n\n"
                "TEL 092-513-9820"
            )
        )

        line_bot_api.reply_message(
        event.reply_token,
        reply
        )
        return
    # =========================
    # 緊急生存確認
    # =========================
    
    elif text == "無事です":

    # USERのメッセージを保存
        save_line_log(
            user,
            "USER",
            "緊急",
            text
            )

        reply = TextSendMessage(
            text=(
                "より安全な場所へ移動してください。\n\n"
                "困っている方がいたら助け合いをお願いします。"
                "こちらとしても安心しますのでGPSを送って下さい"
                "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
                )
            )

    # BOTの返信を保存
        save_line_log(
            user,
            "BOT",
            "緊急",
            "より安全な場所へ移動してください。困っている方がいたら助け合いをお願いします。"
            "こちらとしても安心しますのでGPSを送って下さい"
            "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )
    elif text == "ケガ":

    # USERログ
        save_line_log(
            user,
            "USER",
            "緊急",
            text
            )

        user.step = "wait_gps"
        user.save()

        reply = TextSendMessage(
            text=(
                "ケガをされているとのことですね。\n\n"
                "誰か周りにいたら助けを求めて下さい\n\n"
                "救援や電話での対応ができないか確認いたします。\n\n"
                "まずGPSを送信してください。\n\n"
                "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
                "今の状況が変ったら下記に電話ください\n\n"
                "【委員長】\n"
                "TEL：080-1720-0311\n\n"
                "【組合】\n"
                "TEL：092-513-9820\n\n"
                
                
                )
            )

    # BOTログ
        save_line_log(
            user,
            "BOT",
            "緊急",
            "ケガをされているとのことですね。まず安全を確保してください。GPSを送信してください。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )
        return

    elif text == "危険":

        
        save_line_log(
            user,
            "USER",
            "緊急",
            text
            )

        reply = TextSendMessage(
            text="近くに人はいますか？",
            quick_reply=QuickReply(
                items=[
                    QuickReplyButton(
                        action=MessageAction(
                            label="近くにいます",
                            text="近くにいます"
                            )
                        ),
                    QuickReplyButton(
                        action=MessageAction(
                            label="誰もいません",
                            text="誰もいません"
                            )
                        ),
                    ]
                )
            )
            # BOTログ
        save_line_log(
            user,
            "BOT",
            "緊急",
            "近くに人はいますか？"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )


    elif text == "近くにいます":

    # USERログ

        save_line_log(
            user,
            "USER",
            "緊急",
            text
            )

        reply = TextSendMessage(
            text="近くの人と一緒に安全な場所へ避難してください。\n\n"
            "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
            )

        save_line_log(
            user,
            "BOT",
            "緊急",
            "近くの人と一緒に安全な場所へ避難してください。"
            "こちらとして安心しますのでGPSを送って下さい"
            "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
            )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )


    elif text == "誰もいません":

    # USERログ
        save_line_log(
            user,
            "USER",
            "緊急",
            text
            )

        user.step = "wait_gps"
        user.save()

        reply = TextSendMessage(
            text=(
                "落ち着いてください。\n\n"
                "こちらから救援や電話での対応ができないか確認いたします。\n\n"
                "まずGPSを送ってください。\n"
                "その後、電話番号を入力してください。\n\n"
                "【GPSの送り方】\n\n"
                "① LINEの左下にある「＋」を押します。\n"
                "② 「位置情報」を選択します。\n"
                "③ 「現在地を送信」を押してください。"
                )
            )

    # BOTログ

        save_line_log(
            user,
            "BOT",
            "緊急",
            "落ち着いてください。こちらから救援や電話での対応ができないか確認いたします。まずGPSを送ってください。その後、電話番号を入力してください。"
        )

        line_bot_api.reply_message(
            event.reply_token,
            reply
            )

        return
    
        
@handler.add(MessageEvent, message=LocationMessage)
def handle_location(event):

    user_id = event.source.user_id
    user = LineUser.objects.get(user_id=user_id)

    gps_text = (
    f"緯度: {event.message.latitude}\n"
    f"経度: {event.message.longitude}"
    )

    LineLog.objects.create(
        user=user,
        message_type="location",
        content=gps_text
    )
    

    # 電話番号入力待ち
    user.step = "wait_emergency_phone"
    user.save()

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(
            text="ありがとうございます。\n\n続いて電話番号を入力してください。"
        )
    )
    return

# -------------------------
# LINE LOG
# -------------------------
def line_logs(request):

    logs = (
        LineLog.objects
        .select_related("user")
        .order_by("-created_at")
    )

    return render(
        request,
        "checkapp/line_logs.html",
        {
            "logs": logs,
        }
    )


def export_logs_excel(request):

    wb = Workbook()
    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    blue_fill = PatternFill(
        fill_type="solid",
        fgColor="0066CC"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    ws = wb.active

    ws.title = "LINE LOG"


    # 見出し
    ws.append([
        "日時",
        "名前",
        "地域",
        "送信者",
        "種類",
        "内容"
    ])

    # 見出しを赤にする
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font
    # LINE LOG を取得
    logs = LineLog.objects.select_related("user").order_by("created_at")

    for log in logs:

        ws.append([
            log.created_at.strftime("%Y-%m-%d %H:%M"),
            log.user.name,
            log.user.region,
            log.sender,
            log.message_type,
            log.content
            ])
        print(log.message_type)
        current_row = ws.max_row
        
        content = str(ws.cell(row=current_row, column=6).value)

        # 危険
        if "危険" in content:

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = red_fill
                cell.font = white_font

        # ケガ
        elif "ケガ" in content:

            orange_fill = PatternFill(
                fill_type="solid",
                fgColor="FFA500"
            )

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = orange_fill

        # 電話番号
        elif content.isdigit() and len(content) >= 10:

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = blue_fill
                cell.font = white_font

        # GPS
        # GPS
        elif log.message_type in ["GPS", "location"]:

                gps_fill = PatternFill(
                fill_type="solid",
                fgColor="CCFFFF"
                )

                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = gps_fill


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="LINE_LOG.xlsx"'
    )

    wb.save(response)

    return response

def delete_logs(request):

    LineLog.objects.all().delete()

    return redirect("line_logs")

from .forms import AccountRegisterForm


def account_register(request):

    if request.method == "POST":

        form = AccountRegisterForm(request.POST)

        if form.is_valid():

            Account.objects.create(

                name=form.cleaned_data["name"],
                area=form.cleaned_data["region"],
                login_id=form.cleaned_data["login_id"],
                birth_date=form.cleaned_data["birth_date"],
                password=make_password(
                    form.cleaned_data["password1"]
                ),

            )

            return render(
                request,
                "checkapp/account_register_complete.html",
            )

    else:

        form = AccountRegisterForm()

    return render(
        request,
        "checkapp/account_register.html",
        {
            "form": form,
        },
    )



# =====================================
# ログイン
# =====================================
def login_view(request):

    if request.method == "POST":

        form = LoginForm(request.POST)

        if form.is_valid():

            account = Account.objects.filter(
                login_id=form.cleaned_data["login_id"]
            ).first()

            if account:

                if not account.is_active:

                    form.add_error(
                        None,
                        "このアカウントは利用停止になっています。"
                    )

                elif check_password(
                    form.cleaned_data["password"],
                    account.password,
                ):

                    request.session["account_id"] = account.id
                    request.session["account_name"] = account.name

                    # リッチメニューからの行き先を確認
                    next_page = request.POST.get("next") or request.GET.get("next")

                    if next_page == "news":
                        return redirect("union_news")

                    if next_page == "mycar":
                        return redirect("mycar")
                    
                    if next_page == "roukin":
                        return redirect("roukin")

                    if next_page == "consult":
                        return redirect("consult")

                    if next_page == "broadcast_history":
                        return redirect("broadcast_history")


                    return redirect("union_home")

            form.add_error(
                None,
                "社員番号またはパスワードが違います。"
            )

    else:

        form = LoginForm()

    return render(
        request,
        "checkapp/login.html",
        {
            "form": form,
        }
    )


def logout_view(request):

    request.session.flush()

    return redirect("login")
# =====================================
# アカウント一覧
# =====================================
def account_list(request):

    accounts = Account.objects.all().order_by("login_id")

    return render(
        request,
        "checkapp/account_list.html",
        {
            "accounts": accounts,
        }
    )
    
def account_edit(request, pk):

    account = get_object_or_404(
        Account,
        pk=pk,
    )

    if request.method == "POST":

        account.name = request.POST.get("name")

        account.area = request.POST.get("area")

        account.save()

        return redirect("account_list")

    return render(
        request,
        "checkapp/account_edit.html",
        {
            "account": account,
        },
    )
# =====================================
# 利用停止・利用再開
# =====================================
def account_toggle(request, pk):

    account = get_object_or_404(
        Account,
        pk=pk,
    )

    account.is_active = not account.is_active

    account.save()

    return redirect("account_list")


# =====================================
# アカウント削除
# =====================================
def account_delete(request, pk):

    account = get_object_or_404(
        Account,
        pk=pk,
    )

    account.delete()
    
    return redirect("account_list")
    
def emergency_region(request):

    return render(
        request,
        "checkapp/emergency_region.html"
    )


def emergency_send(request):

    region = request.POST.get("region")

    # 有効な組合員の社員番号だけを取得
    active_login_ids = Account.objects.filter(
        is_active=True
    ).values_list("login_id", flat=True)

    # 選択地域 ＋ 有効Accountと結び付いたLINE利用者だけを対象にする
    users = LineUser.objects.filter(
        region=region,
        login_id__in=active_login_ids
    ).exclude(
        user_id__startswith="web_"
    )

    message = TextSendMessage(
        text=(
            "【緊急生存確認】\n\n"
            "現在の状況を教えてください。"
        ),
        quick_reply=QuickReply(
            items=[
                QuickReplyButton(
                    action=MessageAction(
                        label="無事です",
                        text="無事です"
                    )
                ),
                QuickReplyButton(
                    action=MessageAction(
                        label="ケガ",
                        text="ケガ"
                    )
                ),
                QuickReplyButton(
                    action=MessageAction(
                        label="危険",
                        text="危険"
                    )
                ),
            ]
        )
    )

    success = 0
    error = 0

    for user in users:
        try:
            line_bot_api.push_message(
                user.user_id,
                message
            )
            success += 1

        except Exception as e:
            print("緊急生存確認送信エラー:", e)
            error += 1

    return HttpResponse(
        f"緊急生存確認送信完了 地域:{region} 成功:{success}件 失敗:{error}件"
    )

def clear_test_users(request):

    LineUser.objects.filter(
        user_id__startswith="web_"
    ).delete()

    return redirect("/users/")

def clear_all_users(request):

    LineUser.objects.all().delete()

    return redirect("/users/")
# =====================================
# 組合員専用 PDF表示
# =====================================

def protected_pdf(request, filename):

    # ログインしていなければログイン画面へ
    if "account_id" not in request.session:
        return redirect("login")

    # 開くことを許可するPDF
    allowed_files = {
        "news_05.pdf",
        "news_06.pdf",
        "news_07.pdf",
        "mycar.pdf",
        "roukin.pdf",
    }

    # 許可されていないファイルは開かない
    if filename not in allowed_files:
        raise Http404("PDFが見つかりません")

    pdf_path = Path(settings.BASE_DIR) / "protected_pdfs" / filename

    # PDFが存在しない場合
    if not pdf_path.exists():
        raise Http404("PDFが見つかりません")

    response = FileResponse(
        open(pdf_path, "rb"),
        content_type="application/pdf"
    )
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response
    
def signed_pdf(request, filename):

    # 組合ニュースPDFのみ許可
    allowed_signed_files = {
        "news_05.pdf",
        "news_06.pdf",
        "news_07.pdf",
        "mycar.pdf",
        "roukin.pdf",
    }

    if filename not in allowed_signed_files:
        raise Http404("PDFが見つかりません")

    token = request.GET.get("token")

    if not token:
        raise Http404("無効なPDFリンクです")

    try:
        signed_filename = signing.loads(
            token,
            salt="union-pdf",
            max_age=300
        )
    except signing.BadSignature:
        raise Http404("PDFリンクが無効または期限切れです")

    if signed_filename != filename:
        raise Http404("PDFが見つかりません")

    pdf_path = Path(settings.BASE_DIR) / "protected_pdfs" / filename

    if not pdf_path.exists():
        raise Http404("PDFが見つかりません")

    response = FileResponse(
        open(pdf_path, "rb"),
        content_type="application/pdf"
    )
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


def union_news(request):

    if "account_id" not in request.session:
        return redirect("/login/?next=news")

    news_05_token = signing.dumps(
        "news_05.pdf",
        salt="union-pdf"
    )

    news_06_token = signing.dumps(
        "news_06.pdf",
        salt="union-pdf"
    )

    news_07_token = signing.dumps(
        "news_07.pdf",
        salt="union-pdf"
    )

    return render(
        request,
        "checkapp/news.html",
        {
            "news_05_token": news_05_token,
            "news_06_token": news_06_token,
            "news_07_token": news_07_token,
        }
    )

    
# =====================================
# マイカー共済 専用ページ
# =====================================

def mycar(request):

    if "account_id" not in request.session:
        return redirect("/login/?next=mycar")

    mycar_token = signing.dumps(
        "mycar.pdf",
        salt="union-pdf"
    )

    return render(
        request,
        "checkapp/mycar.html",
        {
            "mycar_token": mycar_token,
        }
    )
def roukin(request):

    # ログインしていなければログイン画面へ
    if "account_id" not in request.session:
        return redirect("/login/?next=roukin")

    roukin_token = signing.dumps(
        "roukin.pdf",
        salt="union-pdf"
    )

    # ログイン済みならNISA・フリーローン専用ページ
    return render(
        request,
        "checkapp/roukin.html",
        {
            "roukin_token": roukin_token,
        }
    )

def consult(request):

    # ログインしていなければログイン画面へ
    if "account_id" not in request.session:
        return redirect("/login/?next=consult")

    # ログイン済みなら相談専用ページ
    return render(
        request,
        "checkapp/consult.html"
    )


def broadcast_history(request):

    # ログインしていなければログイン画面へ
    if "account_id" not in request.session:
        return redirect("/login/?next=broadcast_history")

    # 新しい配信から順番に取得
    broadcasts = BroadcastHistory.objects.all().order_by("-sent_at")

    # PDF付き配信には、過去の配信画面から開くための
    # 署名トークン付きURLを作成する
    for broadcast in broadcasts:
        broadcast.pdf_url = ""

        if broadcast.pdf_filename:
            safe_name = Path(broadcast.pdf_filename).name

            pdf_token = signing.dumps(
                safe_name,
                salt="broadcast-pdf"
            )

            broadcast.pdf_url = (
                f"/broadcast-pdf/{safe_name}/"
                f"?token={pdf_token}"
            )

    return render(
        request,
        "checkapp/broadcast_history.html",
        {
            "broadcasts": broadcasts,
        }
    )


def broadcast_send(request):

    # 入力後、「配信内容を確認する」が押された場合
    if request.method == "POST":

        title = request.POST.get("title", "").strip()
        content = request.POST.get("content", "").strip()
        has_pdf = request.POST.get("has_pdf", "no")
        pdf_file = request.FILES.get("pdf_file")
        pdf_name = ""
        temp_pdf_path = ""

        # PDFありの場合は一時保存
        if has_pdf == "yes" and pdf_file:

            # PDF以外は受け付けない
            if not pdf_file.name.lower().endswith(".pdf"):
                return HttpResponse(
                    "PDFファイルを選択してください。",
                    status=400
                )

            pdf_name = Path(pdf_file.name).name

            temp_pdf_path = default_storage.save(
                f"broadcast_temp/{pdf_name}",
                ContentFile(pdf_file.read())
            )

            # 確認後の配信処理で使えるようセッションへ保存
            request.session["broadcast_temp_pdf"] = temp_pdf_path
            request.session["broadcast_pdf_name"] = pdf_name

        else:
            request.session.pop("broadcast_temp_pdf", None)
            request.session.pop("broadcast_pdf_name", None)

        # 入力内容を確認画面へ渡す
        return render(
            request,
            "checkapp/broadcast_confirm.html",
            {
                "title": title,
                "content": content,
                "has_pdf": has_pdf,
                "pdf_name": pdf_name,
            }
        )

    # 最初は入力画面を表示
    return render(
        request,
        "checkapp/broadcast_send.html"
    )


# =====================================
# 配信用PDF表示
# =====================================
def broadcast_pdf(request, filename):

    safe_filename = Path(filename).name

    # LINE配信用の署名トークンを確認
    token = request.GET.get("token")

    if not token:
        raise Http404("無効なPDFリンクです")

    try:
        signed_filename = signing.loads(
            token,
            salt="broadcast-pdf"
        )
    except signing.BadSignature:
        raise Http404("PDFリンクが無効です")

    if signed_filename != safe_filename:
        raise Http404("PDFが見つかりません")

    pdf_path = (
        Path(settings.BASE_DIR)
        / "protected_pdfs"
        / "broadcasts"
        / safe_filename
    )

    if pdf_path.suffix.lower() != ".pdf":
        raise Http404("PDFが見つかりません")

    if not pdf_path.exists():
        raise Http404("PDFが見つかりません")

    # 案内ページからPDF本体へトークンを引き継ぐ
    return render(
        request,
        "checkapp/broadcast_pdf_open.html",
        {
            "filename": safe_filename,
            "token": token,
        }
    )


# =====================================
# 通常配信・自分だけテスト送信
# =====================================
def broadcast_test_send(request):

    if request.method != "POST":
        return redirect("broadcast_send")

    # テスト送信先はiPhone確認用の社員番号126280だけ
    test_user = LineUser.objects.filter(
        login_id="126280"
    ).exclude(
        user_id__startswith="web_"
    ).first()

    if not test_user:
        return HttpResponse(
            "テスト送信先（社員番号126280）が見つかりません。",
            status=404
        )

    title = request.POST.get("title", "").strip()
    content = request.POST.get("content", "").strip()
    has_pdf = request.POST.get("has_pdf", "no")

    if not title or not content:
        return HttpResponse(
            "タイトルまたは本文がありません。",
            status=400
        )

    message_text = f"【{title}】\n\n{content}"

    # PDFありの場合
    if has_pdf == "yes":

        temp_pdf_path = request.session.get(
            "broadcast_temp_pdf"
        )
        pdf_name = request.session.get(
            "broadcast_pdf_name"
        )

        if not temp_pdf_path or not pdf_name:
            return HttpResponse(
                "PDFが見つかりません。入力画面からやり直してください。",
                status=400
            )

        # 一時保存PDFの実際の場所
        source_path = Path(
            default_storage.path(temp_pdf_path)
        )

        if not source_path.exists():
            return HttpResponse(
                "一時保存したPDFが見つかりません。",
                status=404
            )

        # 配信用PDFフォルダへ正式保存
        safe_name = Path(pdf_name).name

        destination_path = (
            Path(settings.BASE_DIR)
            / "protected_pdfs"
            / "broadcasts"
            / safe_name
        )

        destination_path.write_bytes(
            source_path.read_bytes()
        )

        # 配信PDF専用の署名トークンを作る
        pdf_token = signing.dumps(
            safe_name,
            salt="broadcast-pdf"
        )

        # トークン付きPDF案内ページURLを作る
        # LINEから開ける公開URLを使用
        public_base_url = "https://nonfrigid-smug-candance.ngrok-free.dev"

        pdf_url = (
            f"{public_base_url}/broadcast-pdf/"
            f"{safe_name}/?token={pdf_token}"
        )

        message_text += (
            "\n\n▼ PDFはこちら\n"
            + pdf_url
        )

    # 126280だけへ送信
    try:
        line_bot_api.push_message(
            test_user.user_id,
            TextSendMessage(text=message_text)
        )

    except Exception as e:
        print("通常配信テスト送信エラー:", e)

        return HttpResponse(
            "LINEへのテスト送信に失敗しました。",
            status=500
        )

    return HttpResponse(
        "社員番号126280へのテスト配信が完了しました。"
    )


# =====================================
# 配信用PDF本体
# =====================================
def broadcast_pdf_file(request, filename):

    safe_filename = Path(filename).name

    # LINE配信用の署名トークンを確認
    token = request.GET.get("token")

    if not token:
        raise Http404("無効なPDFリンクです")

    try:
        signed_filename = signing.loads(
            token,
            salt="broadcast-pdf"
        )
    except signing.BadSignature:
        raise Http404("PDFリンクが無効です")

    if signed_filename != safe_filename:
        raise Http404("PDFが見つかりません")

    pdf_path = (
        Path(settings.BASE_DIR)
        / "protected_pdfs"
        / "broadcasts"
        / safe_filename
    )

    if pdf_path.suffix.lower() != ".pdf":
        raise Http404("PDFが見つかりません")

    if not pdf_path.exists():
        raise Http404("PDFが見つかりません")

    response = FileResponse(
        open(pdf_path, "rb"),
        content_type="application/pdf"
    )

    # Android対策：PDF本体はダウンロード方式
    response["Content-Disposition"] = (
        f'attachment; filename="document.pdf"'
    )

    return response


# =====================================
# お知らせ・正式全組合員配信
# =====================================
def broadcast_all_send(request):

    if request.method != "POST":
        return redirect("broadcast_send")

    title = request.POST.get("title", "").strip()
    content = request.POST.get("content", "").strip()
    has_pdf = request.POST.get("has_pdf", "no")

    if not title or not content:
        return HttpResponse(
            "タイトルまたは本文がありません。",
            status=400
        )

    # 現在有効な組合員の社員番号だけを取得
    active_login_ids = Account.objects.filter(
        is_active=True
    ).values_list(
        "login_id",
        flat=True
    )

    # 有効組合員と社員番号が一致するLINE登録者だけを対象にする
    users = LineUser.objects.filter(
        login_id__in=active_login_ids
    ).exclude(
        user_id__startswith="web_"
    )

    if not users.exists():
        return HttpResponse(
            "配信対象の組合員がいません。",
            status=404
        )

    message_text = f"【{title}】\n\n{content}"

    # PDFありの場合
    if has_pdf == "yes":

        temp_pdf_path = request.session.get(
            "broadcast_temp_pdf"
        )
        pdf_name = request.session.get(
            "broadcast_pdf_name"
        )

        if not temp_pdf_path or not pdf_name:
            return HttpResponse(
                "PDFが見つかりません。入力画面からやり直してください。",
                status=400
            )

        source_path = Path(
            default_storage.path(temp_pdf_path)
        )

        if not source_path.exists():
            return HttpResponse(
                "一時保存したPDFが見つかりません。",
                status=404
            )

        safe_name = Path(pdf_name).name

        destination_path = (
            Path(settings.BASE_DIR)
            / "protected_pdfs"
            / "broadcasts"
            / safe_name
        )

        destination_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        destination_path.write_bytes(
            source_path.read_bytes()
        )

        # Android・iPhoneで確認済みの署名トークン方式
        pdf_token = signing.dumps(
            safe_name,
            salt="broadcast-pdf"
        )

        # LINEから開ける公開URLを使用
        public_base_url = "https://nonfrigid-smug-candance.ngrok-free.dev"

        pdf_url = (
            f"{public_base_url}/broadcast-pdf/"
            f"{safe_name}/?token={pdf_token}"
        )

        message_text += (
            "\n\n▼ PDFはこちら\n"
            + pdf_url
        )

    success = 0
    error = 0

    # 有効組合員だけへ配信
    for user in users:

        try:
            line_bot_api.push_message(
                user.user_id,
                TextSendMessage(text=message_text)
            )
            success += 1

        except Exception as e:
            print(
                "お知らせ配信エラー:",
                user.login_id,
                e
            )
            error += 1

    # 正式配信の履歴を保存
    BroadcastHistory.objects.create(
        title=title,
        content=content,
        pdf_filename=pdf_name if has_pdf == "yes" else ""
    )

    return HttpResponse(
        f"全組合員への配信が完了しました。"
        f" 成功：{success}件 / 失敗：{error}件"
    )


# =====================================
# お知らせ・全組合員配信 最終確認
# =====================================
def broadcast_all_confirm(request):

    if request.method != "POST":
        return redirect("broadcast_send")

    title = request.POST.get("title", "").strip()
    content = request.POST.get("content", "").strip()
    has_pdf = request.POST.get("has_pdf", "no")

    if not title or not content:
        return HttpResponse(
            "タイトルまたは本文がありません。",
            status=400
        )

    # 現在有効な組合員
    active_login_ids = Account.objects.filter(
        is_active=True
    ).values_list(
        "login_id",
        flat=True
    )

    # 実際にLINE配信できる対象者
    users = LineUser.objects.filter(
        login_id__in=active_login_ids
    ).exclude(
        user_id__startswith="web_"
    )

    target_count = users.count()

    # PDF情報はセッションに保存済みのものを使用
    pdf_name = ""

    if has_pdf == "yes":
        pdf_name = request.session.get(
            "broadcast_pdf_name",
            ""
        )

        if not pdf_name:
            return HttpResponse(
                "PDFが見つかりません。入力画面からやり直してください。",
                status=400
            )

    return render(
        request,
        "checkapp/broadcast_all_confirm.html",
        {
            "title": title,
            "content": content,
            "has_pdf": has_pdf,
            "pdf_name": pdf_name,
            "target_count": target_count,
        }
    )
